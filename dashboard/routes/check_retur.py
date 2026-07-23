"""
dashboard/routes/check_retur.py
Fitur "Check Retur" — pencatatan hasil pengecekan barang retur.

Semua orang yang login boleh INPUT data baru.
Cuma yang punya permission "edit_check_retur" (atau superadmin) yang boleh
EDIT / HAPUS data yang udah masuk — biar histori pengecekan gak sembarangan
diubah orang.

Master data "Kode Artikel" + "Nama Artikel" dikelola terpisah oleh superadmin
di /superadmin/artikel (lihat routes/superadmin.py). Form di sini cuma
"mengonsumsi" master data itu buat dropdown + auto-fill kode.
"""

import os
import io
import uuid
from datetime import datetime, timezone, timedelta
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    send_from_directory, send_file, abort,
)
from werkzeug.utils import secure_filename
from dashboard.auth import login_required, get_current_user, has_permission
from db.local import get_conn

check_retur_bp = Blueprint("check_retur", __name__, url_prefix="/check-retur")

WIB = timezone(timedelta(hours=7))

# Foto disimpan di folder yang sama dengan volume "data-internal" yang udah
# di-mount di docker-compose.yml, biar persist walau container di-restart.
BASE_PATH = os.environ.get("STORAGE_BASE_PATH", "/app/data-internal")
FOTO_DIR = os.path.join(BASE_PATH, "check-retur-photos")

ALLOWED_FOTO_EXT = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
KONDISI_VALID = ("waste", "ok", "service")


def _now_wib_str() -> str:
    return datetime.now(WIB).strftime("%Y-%m-%d %H:%M:%S")


def _simpan_foto(file_storage) -> str | None:
    """Simpan foto upload (kalau ada), return nama file yang disimpan atau None."""
    if not file_storage or not file_storage.filename:
        return None
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in ALLOWED_FOTO_EXT:
        return None
    os.makedirs(FOTO_DIR, exist_ok=True)
    nama_aman = secure_filename(file_storage.filename)
    nama_file = f"{uuid.uuid4().hex}_{nama_aman}"
    file_storage.save(os.path.join(FOTO_DIR, nama_file))
    return nama_file


def _hapus_foto(nama_file: str | None):
    if not nama_file:
        return
    path = os.path.join(FOTO_DIR, nama_file)
    if os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass


def _ambil_filter_args():
    return {
        "q": request.args.get("q", "").strip(),
        "artikel": request.args.get("artikel", "").strip(),
        "kondisi": request.args.get("kondisi", "").strip(),
    }


def _query_rows(q: str, artikel: str, kondisi: str, limit: int | None = 200):
    """Query check_retur dengan filter pencarian teks + dropdown artikel/kondisi.
    Dipakai bareng sama halaman list & export Excel biar hasilnya konsisten."""
    where = []
    params = []

    if q:
        like = f"%{q}%"
        where.append("(cr.no_surat ILIKE ? OR cr.nama_artikel ILIKE ? OR cr.kode_artikel::text ILIKE ? OR cr.serial_number ILIKE ?)")
        params += [like, like, like, like]
    if artikel:
        where.append("cr.nama_artikel = ?")
        params.append(artikel)
    if kondisi in KONDISI_VALID:
        where.append("cr.kondisi = ?")
        params.append(kondisi)

    sql = """
        SELECT cr.*, w.nama as dicek_oleh_nama
        FROM check_retur cr
        LEFT JOIN whitelist w ON cr.dicek_oleh = w.user_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY cr.created_at DESC"
    if limit:
        sql += f" LIMIT {int(limit)}"

    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


# ──────────────────────────────────────────
# List + input
# ──────────────────────────────────────────

@check_retur_bp.route("/")
@login_required
def index():
    user = get_current_user()
    bisa_edit = has_permission("edit_check_retur")
    f = _ambil_filter_args()

    rows = _query_rows(f["q"], f["artikel"], f["kondisi"])

    with get_conn() as conn:
        artikel_list = conn.execute("SELECT kode, nama FROM artikel ORDER BY nama").fetchall()

    return render_template("check_retur.html",
        user=user, rows=rows, bisa_edit=bisa_edit,
        artikel_list=artikel_list, **f,
    )


@check_retur_bp.route("/download/excel")
@login_required
def download_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    f = _ambil_filter_args()
    rows = _query_rows(f["q"], f["artikel"], f["kondisi"], limit=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Check Retur"

    # Sengaja TANPA kolom foto -- biar file Excel-nya kecil & ringan.
    headers = ["No Surat", "Nama Artikel", "Kode Artikel", "Serial Number",
               "Kondisi", "Keterangan", "Dicek Oleh", "Tanggal"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for r in rows:
        ws.append([
            r["no_surat"] or "",
            r["nama_artikel"],
            r["kode_artikel"],
            r["serial_number"] or "",
            (r["kondisi"] or "").upper(),
            r["keterangan"] or "",
            r["dicek_oleh_nama"] or "",
            r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
        ])

    lebar = [16, 26, 12, 16, 10, 34, 18, 17]
    for i, w in enumerate(lebar, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nama_file = f"check-retur-{datetime.now(WIB).strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nama_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@check_retur_bp.route("/buat", methods=["POST"])
@login_required
def buat():
    user = get_current_user()
    no_surat = request.form.get("no_surat", "").strip()
    nama_artikel = request.form.get("nama_artikel", "").strip()
    kode_artikel_str = request.form.get("kode_artikel", "").strip()
    serial_number = request.form.get("serial_number", "").strip()
    kondisi = request.form.get("kondisi", "").strip()
    keterangan = request.form.get("keterangan", "").strip()
    foto = request.files.get("foto")

    # No Surat sekarang opsional -- gak semua barang retur ada surat jalannya.
    if not (nama_artikel and kondisi in KONDISI_VALID):
        return redirect(url_for("check_retur.index"))

    kode_artikel = int(kode_artikel_str) if kode_artikel_str.isdigit() else None
    foto_path = _simpan_foto(foto)

    with get_conn() as conn:
        conn.execute("""
            INSERT INTO check_retur
                (no_surat, nama_artikel, kode_artikel, serial_number, kondisi, foto_path, keterangan, dicek_oleh)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (no_surat, nama_artikel, kode_artikel, serial_number or None, kondisi,
              foto_path, keterangan or None, user["user_id"]))
        conn.commit()

    return redirect(url_for("check_retur.index"))


@check_retur_bp.route("/<int:cr_id>/edit", methods=["POST"])
@login_required
def edit(cr_id):
    if not has_permission("edit_check_retur"):
        abort(403)

    no_surat = request.form.get("no_surat", "").strip()
    nama_artikel = request.form.get("nama_artikel", "").strip()
    kode_artikel_str = request.form.get("kode_artikel", "").strip()
    serial_number = request.form.get("serial_number", "").strip()
    kondisi = request.form.get("kondisi", "").strip()
    keterangan = request.form.get("keterangan", "").strip()
    hapus_foto_lama = request.form.get("hapus_foto") == "on"
    foto = request.files.get("foto")

    if not (nama_artikel and kondisi in KONDISI_VALID):
        return redirect(url_for("check_retur.index"))

    kode_artikel = int(kode_artikel_str) if kode_artikel_str.isdigit() else None

    with get_conn() as conn:
        row = conn.execute("SELECT foto_path FROM check_retur WHERE id = ?", (cr_id,)).fetchone()
        if not row:
            return redirect(url_for("check_retur.index"))

        foto_path = row["foto_path"]
        foto_baru = _simpan_foto(foto)
        if foto_baru:
            _hapus_foto(foto_path)
            foto_path = foto_baru
        elif hapus_foto_lama:
            _hapus_foto(foto_path)
            foto_path = None

        conn.execute("""
            UPDATE check_retur SET
                no_surat = ?, nama_artikel = ?, kode_artikel = ?, serial_number = ?, kondisi = ?,
                foto_path = ?, keterangan = ?, updated_at = ?
            WHERE id = ?
        """, (no_surat, nama_artikel, kode_artikel, serial_number or None, kondisi,
              foto_path, keterangan or None, _now_wib_str(), cr_id))
        conn.commit()

    return redirect(url_for("check_retur.index"))


@check_retur_bp.route("/<int:cr_id>/hapus", methods=["POST"])
@login_required
def hapus(cr_id):
    if not has_permission("edit_check_retur"):
        abort(403)
    with get_conn() as conn:
        row = conn.execute("SELECT foto_path FROM check_retur WHERE id = ?", (cr_id,)).fetchone()
        if row:
            _hapus_foto(row["foto_path"])
        conn.execute("DELETE FROM check_retur WHERE id = ?", (cr_id,))
        conn.commit()
    return redirect(url_for("check_retur.index"))


@check_retur_bp.route("/foto/<path:filename>")
@login_required
def foto(filename):
    return send_from_directory(FOTO_DIR, filename)