"""
dashboard/routes/jadwal.py
Route manajemen jadwal
"""

import io
import random
from datetime import datetime, date, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file
from dashboard.auth import login_required, get_current_user, has_permission
from db.local import get_conn, upsert_jadwal, delete_jadwal, get_jadwal_by_bulan, get_all_nama_jadwal

jadwal_bp = Blueprint("jadwal", __name__, url_prefix="/jadwal")

WIB = timezone(timedelta(hours=7))

BULAN_NAMA = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
              "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

# Nama yang cuma boleh piket maksimal 1x sebulan
PIKET_LIMIT_KHUSUS = {
    "Azmii": 1,
    "Danu": 1,
    "Cumey": 1,
}
PIKET_LIBUR_KERJA_ONLY = {"Cumey"}


def can_edit() -> bool:
    """Boleh edit jadwal (upload/tambah/hapus/generate/blackout) kalau punya permission ini."""
    return has_permission("generate_jadwal")


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def get_hari_libur(tahun: int) -> dict:
    """Ambil hari libur dari database, key-nya string YYYY-MM-DD."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT tanggal, nama FROM hari_libur WHERE EXTRACT(YEAR FROM tanggal) = ?",
            (tahun,)
        ).fetchall()
        return {r["tanggal"].isoformat(): r["nama"] for r in rows}


# ──────────────────────────────────────────
# Kalender
# ──────────────────────────────────────────
@jadwal_bp.route("/")
@login_required
def index():
    user = get_current_user()
    tahun = int(request.args.get("tahun", date.today().year))
    bulan = int(request.args.get("bulan", date.today().month))

    jadwal_list = get_jadwal_by_bulan(tahun, bulan)
    hari_libur = get_hari_libur(tahun)

    jadwal_map = {}
    for j in jadwal_list:
        tgl = j["tanggal"]
        if tgl not in jadwal_map:
            jadwal_map[tgl] = {"oc": [], "piket": [], "off": []}
        jadwal_map[tgl][j["tipe"]].append(j)

    import calendar
    _, days_in_month = calendar.monthrange(tahun, bulan)
    first_weekday = date(tahun, bulan, 1).weekday()

    nama_list = get_all_nama_jadwal()
    awal_bulan = f"{tahun}-{bulan:02d}-01"
    with get_conn() as conn:
        whitelist_all = conn.execute(
            "SELECT user_id, nama FROM whitelist ORDER BY nama"
        ).fetchall()
        blackout_list = conn.execute("""
            SELECT b.*, w.nama as nama_user
            FROM blackout b
            JOIN whitelist w ON b.whitelist_id = w.user_id
            WHERE b.tanggal >= ? AND b.tanggal < (?::date + INTERVAL '1 month')
            ORDER BY b.tanggal, w.nama
        """, (awal_bulan, awal_bulan)).fetchall()

    return render_template("jadwal.html",
        user=user,
        tahun=tahun,
        bulan=bulan,
        bulan_nama=BULAN_NAMA[bulan - 1],
        days_in_month=days_in_month,
        first_weekday=first_weekday,
        jadwal_map=jadwal_map,
        hari_libur=hari_libur,
        nama_list=nama_list,
        can_edit=can_edit(),
        today_str=date.today().strftime("%Y-%m-%d"),
        whitelist_all=whitelist_all,
        blackout_list=blackout_list,
    )


# ──────────────────────────────────────────
# Upload Excel
# ──────────────────────────────────────────
@jadwal_bp.route("/upload", methods=["POST"])
@login_required
def upload():
    if not can_edit():
        return redirect(url_for("jadwal.index"))

    file = request.files.get("file")
    if not file or not file.filename.endswith(".xlsx"):
        return redirect(url_for("jadwal.index"))

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    try:
        idx_nama = headers.index("nama")
        idx_tanggal = headers.index("tanggal")
        idx_oc = headers.index("oc")
        idx_piket = headers.index("piket")
        idx_off = headers.index("off")
    except ValueError:
        return redirect(url_for("jadwal.index"))

    for row in ws.iter_rows(min_row=2, values_only=True):
        nama = str(row[idx_nama]).strip() if row[idx_nama] else None
        tanggal = row[idx_tanggal]
        oc, piket, off = row[idx_oc], row[idx_piket], row[idx_off]

        if not nama or not tanggal:
            continue

        tanggal_str = tanggal.strftime("%Y-%m-%d") if isinstance(tanggal, datetime) else str(tanggal).strip()

        if oc and str(oc).upper() in ("Y", "YES", "1", "TRUE", "✓"):
            upsert_jadwal(nama, tanggal_str, "oc")
        if piket and str(piket).upper() in ("Y", "YES", "1", "TRUE", "✓"):
            upsert_jadwal(nama, tanggal_str, "piket")
        if off and str(off).upper() in ("Y", "YES", "1", "TRUE", "✓"):
            upsert_jadwal(nama, tanggal_str, "off")

    return redirect(url_for("jadwal.index"))


# ──────────────────────────────────────────
# Input manual
# ──────────────────────────────────────────
@jadwal_bp.route("/tambah", methods=["POST"])
@login_required
def tambah():
    if not can_edit():
        return redirect(url_for("jadwal.index"))

    nama = request.form.get("nama", "").strip()
    tanggal = request.form.get("tanggal", "").strip()
    tipe = request.form.get("tipe", "").strip()

    if nama and tanggal and tipe in ("oc", "piket", "off"):
        upsert_jadwal(nama, tanggal, tipe)

    return redirect(url_for("jadwal.index") + f"?tahun={tanggal[:4]}&bulan={int(tanggal[5:7])}")


@jadwal_bp.route("/hapus/<int:jadwal_id>", methods=["POST"])
@login_required
def hapus(jadwal_id):
    if not can_edit():
        return redirect(url_for("jadwal.index"))
    delete_jadwal(jadwal_id)
    return redirect(url_for("jadwal.index"))


# ──────────────────────────────────────────
# Template Excel
# ──────────────────────────────────────────
@jadwal_bp.route("/template")
@login_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jadwal"

    headers = ["nama", "tanggal", "oc", "piket", "off"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    contoh = [
        ["Azmii", "2026-07-01", "Y", "", ""],
        ["Noval", "2026-07-01", "", "Y", ""],
        ["Febry", "2026-07-02", "", "", "Y"],
    ]
    for row_data in contoh:
        ws.append(row_data)

    for col, w in zip("ABCDE", [20, 15, 8, 8, 8]):
        ws.column_dimensions[col].width = w

    ws.append([])
    ws.append(["Keterangan: isi kolom OC/Piket/Off dengan Y jika bertugas"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name="template_jadwal.xlsx")


# ──────────────────────────────────────────
# Blackout (Request Tanggal)
# ──────────────────────────────────────────
@jadwal_bp.route("/blackout/tambah", methods=["POST"])
@login_required
def tambah_blackout():
    if not can_edit():
        return redirect(url_for("jadwal.index"))

    whitelist_id = request.form.get("whitelist_id", "").strip()
    tanggal_mulai = request.form.get("tanggal_mulai", "").strip()
    tanggal_selesai = request.form.get("tanggal_selesai", "").strip()
    keterangan = request.form.get("keterangan", "").strip()
    tahun = request.form.get("tahun", date.today().year)
    bulan = request.form.get("bulan", date.today().month)

    if whitelist_id.isdigit() and tanggal_mulai:
        start = datetime.strptime(tanggal_mulai, "%Y-%m-%d").date()
        end = datetime.strptime(tanggal_selesai, "%Y-%m-%d").date() if tanggal_selesai else start
        with get_conn() as conn:
            current = start
            while current <= end:
                conn.execute("""
                    INSERT INTO blackout (whitelist_id, tanggal, keterangan, dibuat_oleh)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT (whitelist_id, tanggal) DO NOTHING
                """, (int(whitelist_id), current.strftime("%Y-%m-%d"), keterangan or None, session.get("user_id")))
                current += timedelta(days=1)
            conn.commit()

    return redirect(url_for("jadwal.index", tahun=tahun, bulan=bulan))


@jadwal_bp.route("/blackout/hapus/<int:bid>", methods=["POST"])
@login_required
def hapus_blackout(bid):
    if not can_edit():
        return redirect(url_for("jadwal.index"))
    tahun = request.form.get("tahun", date.today().year)
    bulan = request.form.get("bulan", date.today().month)
    with get_conn() as conn:
        conn.execute("DELETE FROM blackout WHERE id = ?", (bid,))
        conn.commit()
    return redirect(url_for("jadwal.index", tahun=tahun, bulan=bulan))


# ──────────────────────────────────────────
# Generate Jadwal
# ──────────────────────────────────────────
def _generate_jadwal(tahun: int, bulan: int) -> dict:
    """Logic generate jadwal — random tapi tetap merata, dengan limit & aturan khusus."""
    import calendar as cal_module

    with get_conn() as conn:
        oc_list = conn.execute(
            "SELECT user_id, nama FROM whitelist ORDER BY added_at, user_id"
        ).fetchall()
        piket_list = conn.execute("""
            SELECT dp.whitelist_id as user_id, w.nama
            FROM daftar_piket dp
            JOIN whitelist w ON dp.whitelist_id = w.user_id
            ORDER BY dp.urutan
        """).fetchall()

    oc_ids = [r["user_id"] for r in oc_list]
    oc_names = {r["user_id"]: r["nama"] for r in oc_list}
    piket_ids = [r["user_id"] for r in piket_list]
    piket_names = {r["user_id"]: r["nama"] for r in piket_list}

    if not oc_ids:
        return {}

    hari_libur = get_hari_libur(tahun)

    awal_bulan = f"{tahun}-{bulan:02d}-01"
    with get_conn() as conn:
        bo_rows = conn.execute("""
            SELECT whitelist_id, tanggal FROM blackout
            WHERE tanggal >= ? AND tanggal < (?::date + INTERVAL '1 month')
        """, (awal_bulan, awal_bulan)).fetchall()
    blackout = {}
    for r in bo_rows:
        blackout.setdefault(r["whitelist_id"], set()).add(r["tanggal"].isoformat())

    _, days = cal_module.monthrange(tahun, bulan)
    result = {}
    oc_count = {uid: 0 for uid in oc_ids}
    piket_count = {uid: 0 for uid in piket_ids}

    def limit_piket_untuk(uid):
        nama = piket_names.get(uid)
        return PIKET_LIMIT_KHUSUS.get(nama)

    def boleh_piket_hari_ini(uid, is_weekend, is_libur):
        nama = piket_names.get(uid)
        if nama in PIKET_LIBUR_KERJA_ONLY:
            # cuma boleh piket kalau HARI KERJA (Senin-Jumat) DAN tanggal merah
            return (not is_weekend) and is_libur
        return True

    for day in range(1, days + 1):
        tgl = f"{tahun}-{bulan:02d}-{day:02d}"
        dow = date(tahun, bulan, day).weekday()
        is_weekend = dow >= 5
        is_libur = tgl in hari_libur
        need_piket = is_weekend or is_libur

        result[tgl] = {
            "oc": None, "oc_wid": None, "oc_merah": False,
            "piket": None, "piket_wid": None, "piket_merah": False
        }

        piket_today = None
        if need_piket and piket_ids:
            kandidat = [
                uid for uid in piket_ids
                if tgl not in blackout.get(uid, set())
                and (limit_piket_untuk(uid) is None or piket_count[uid] < limit_piket_untuk(uid))
                and boleh_piket_hari_ini(uid, is_weekend, is_libur)
            ]
            if kandidat:
                min_count = min(piket_count[uid] for uid in kandidat)
                paling_sedikit = [uid for uid in kandidat if piket_count[uid] == min_count]
                pilih = random.choice(paling_sedikit)
                result[tgl]["piket"] = piket_names[pilih]
                result[tgl]["piket_wid"] = pilih
                piket_count[pilih] += 1
                piket_today = pilih
            else:
                result[tgl]["piket_merah"] = True

        if oc_ids:
            kandidat = [
                uid for uid in oc_ids
                if tgl not in blackout.get(uid, set())
                and uid != piket_today
            ]
            if kandidat:
                min_count = min(oc_count[uid] for uid in kandidat)
                paling_sedikit = [uid for uid in kandidat if oc_count[uid] == min_count]
                pilih = random.choice(paling_sedikit)
                result[tgl]["oc"] = oc_names[pilih]
                result[tgl]["oc_wid"] = pilih
                oc_count[pilih] += 1
            else:
                result[tgl]["oc_merah"] = True

    return result


def _update_rolling_state(tahun: int, bulan: int, hasil: dict):
    last_oc = last_piket = None
    for tgl in sorted(hasil.keys(), reverse=True):
        if hasil[tgl]["oc_wid"] and not last_oc:
            last_oc = hasil[tgl]["oc_wid"]
        if hasil[tgl]["piket_wid"] and not last_piket:
            last_piket = hasil[tgl]["piket_wid"]
        if last_oc and last_piket:
            break
    now_str = datetime.now(WIB).strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        if last_oc:
            conn.execute(
                "UPDATE rolling_state SET last_whitelist_id = ?, updated_at = ? WHERE tipe = 'oc'",
                (last_oc, now_str)
            )
        if last_piket:
            conn.execute(
                "UPDATE rolling_state SET last_whitelist_id = ?, updated_at = ? WHERE tipe = 'piket'",
                (last_piket, now_str)
            )
        conn.commit()


@jadwal_bp.route("/preview", methods=["POST"])
@login_required
def preview():
    if not can_edit():
        return redirect(url_for("jadwal.index"))

    tahun = int(request.form.get("tahun", date.today().year))
    bulan = int(request.form.get("bulan", date.today().month))
    hasil = _generate_jadwal(tahun, bulan)

    with get_conn() as conn:
        whitelist = conn.execute("SELECT user_id, nama, no_hp FROM whitelist ORDER BY nama").fetchall()

    ringkasan = {}
    for w in whitelist:
        ringkasan[w["nama"]] = {"no_hp": w["no_hp"] or "-", "oc": [], "oc_merah": [], "piket": [], "piket_merah": []}

    for tgl, data in sorted(hasil.items()):
        day = int(tgl.split("-")[2])
        if data["oc"] and data["oc"] in ringkasan:
            (ringkasan[data["oc"]]["oc_merah"] if data["oc_merah"] else ringkasan[data["oc"]]["oc"]).append(str(day))
        if data["piket"] and data["piket"] in ringkasan:
            (ringkasan[data["piket"]]["piket_merah"] if data["piket_merah"] else ringkasan[data["piket"]]["piket"]).append(str(day))

    return render_template("jadwal_preview.html",
        user=get_current_user(), tahun=tahun, bulan=bulan, bulan_nama=BULAN_NAMA[bulan - 1],
        ringkasan=ringkasan, hasil=hasil,
    )


@jadwal_bp.route("/simpan", methods=["POST"])
@login_required
def simpan_generate():
    if not can_edit():
        return redirect(url_for("jadwal.index"))

    tahun = int(request.form.get("tahun", date.today().year))
    bulan = int(request.form.get("bulan", date.today().month))
    hasil = _generate_jadwal(tahun, bulan)
    awal_bulan = f"{tahun}-{bulan:02d}-01"

    with get_conn() as conn:
        conn.execute(
            "DELETE FROM jadwal WHERE tanggal >= ? AND tanggal < (?::date + INTERVAL '1 month')",
            (awal_bulan, awal_bulan)
        )
        for tgl, data in hasil.items():
            if data["oc"] and not data["oc_merah"]:
                conn.execute(
                    "INSERT INTO jadwal (nama, tanggal, tipe) VALUES (?, ?, 'oc') ON CONFLICT (nama, tanggal, tipe) DO NOTHING",
                    (data["oc"], tgl)
                )
            if data["piket"] and not data["piket_merah"]:
                conn.execute(
                    "INSERT INTO jadwal (nama, tanggal, tipe) VALUES (?, ?, 'piket') ON CONFLICT (nama, tanggal, tipe) DO NOTHING",
                    (data["piket"], tgl)
                )
        conn.commit()

    _update_rolling_state(tahun, bulan, hasil)
    return redirect(url_for("jadwal.index", tahun=tahun, bulan=bulan))


# ──────────────────────────────────────────
# Download — Excel & Gambar (dua-duanya kebuka semua orang, baca dari data TERSIMPAN)
# ──────────────────────────────────────────
def _build_ringkasan(tahun: int, bulan: int):
    """
    Susun ringkasan per-nama dari data jadwal yang UDAH TERSIMPAN (bukan generate baru).
    Pencocokan nama case-insensitive + trimmed, biar typo dikit (kapital/spasi) gak bikin
    orang itu ke-skip diem-diem. Nama yang tetep gak ketemu di whitelist dikumpulin di `unmatched`.
    """
    awal_bulan = f"{tahun}-{bulan:02d}-01"
    with get_conn() as conn:
        whitelist = conn.execute("SELECT user_id, nama, no_hp FROM whitelist ORDER BY nama").fetchall()
        jadwal_list = conn.execute("""
            SELECT nama, tanggal, tipe FROM jadwal
            WHERE tanggal >= ? AND tanggal < (?::date + INTERVAL '1 month')
            ORDER BY tanggal
        """, (awal_bulan, awal_bulan)).fetchall()

    nama_lookup = {_norm(w["nama"]): w["nama"] for w in whitelist}
    ringkasan = {w["nama"]: {"no_hp": w["no_hp"] or "-", "oc": [], "piket": []} for w in whitelist}

    unmatched = set()
    for j in jadwal_list:
        nama_asli = nama_lookup.get(_norm(j["nama"]))
        if not nama_asli:
            unmatched.add(j["nama"])
            continue
        day = j["tanggal"].day
        if j["tipe"] == "oc":
            ringkasan[nama_asli]["oc"].append(str(day))
        elif j["tipe"] == "piket":
            ringkasan[nama_asli]["piket"].append(str(day))

    return ringkasan, sorted(unmatched)


@jadwal_bp.route("/download/excel")
@login_required
def download_excel():
    tahun = int(request.args.get("tahun", date.today().year))
    bulan = int(request.args.get("bulan", date.today().month))
    bulan_nama = BULAN_NAMA[bulan - 1]
    ringkasan, unmatched = _build_ringkasan(tahun, bulan)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jadwal {bulan_nama} {tahun}"

    blue_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:D1")
    ws["A1"] = f"JADWAL ON CALL MALAM & PIKET — {bulan_nama.upper()} {tahun}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ["NAMA", "NO HANDPHONE", "TANGGAL ON CALL MALAM",
               "TANGGAL PIKET WEEKEND\nDAN HARI LIBUR MASUK KANTOR"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = blue_fill, header_font, center, thin
    ws.row_dimensions[2].height = 40

    row_idx = 3
    for nama, data in ringkasan.items():
        ws.cell(row=row_idx, column=1, value=nama).border = thin
        ws.cell(row=row_idx, column=2, value=data["no_hp"]).border = thin
        oc_cell = ws.cell(row=row_idx, column=3, value=", ".join(data["oc"]) if data["oc"] else "-")
        oc_cell.border, oc_cell.alignment = thin, center
        piket_cell = ws.cell(row=row_idx, column=4, value=", ".join(data["piket"]) if data["piket"] else "-")
        piket_cell.border, piket_cell.alignment = thin, center
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    if unmatched:
        row_idx += 1
        ws.merge_cells(f"A{row_idx}:D{row_idx}")
        cell = ws.cell(row=row_idx, column=1,
                        value=f"⚠️ Nama di jadwal gak ketemu di whitelist (cek typo): {', '.join(unmatched)}")
        cell.font = Font(color="B91C1C", italic=True, size=10)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                      as_attachment=True, download_name=f"jadwal_{bulan_nama}_{tahun}.xlsx")


def _font(size, bold=False):
    from PIL import ImageFont
    path = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    try:
        return ImageFont.truetype(path, size)
    except Exception:
        return ImageFont.load_default()


@jadwal_bp.route("/download/gambar")
@login_required
def download_gambar():
    from PIL import Image, ImageDraw

    tahun = int(request.args.get("tahun", date.today().year))
    bulan = int(request.args.get("bulan", date.today().month))
    bulan_nama = BULAN_NAMA[bulan - 1]
    ringkasan, unmatched = _build_ringkasan(tahun, bulan)

    col_w = [170, 140, 260, 300]
    row_h = 32
    header_h = 60
    top_pad = 20
    n_rows = len(ringkasan)
    warn_h = 26 if unmatched else 0
    width = sum(col_w) + 40
    height = top_pad + header_h + row_h * n_rows + warn_h + 30

    img = Image.new("RGB", (width, max(height, 200)), "white")
    draw = ImageDraw.Draw(img)

    title = f"JADWAL ON CALL MALAM & PIKET — {bulan_nama.upper()} {tahun}"
    draw.text((20, top_pad), title, fill="#1F4E79", font=_font(17, bold=True))

    y = top_pad + 34
    draw.rectangle([20, y, width - 20, y + header_h - 10], fill="#1F4E79")
    x = 20
    for i, h in enumerate(["NAMA", "NO HP", "TANGGAL OC", "TANGGAL PIKET"]):
        draw.text((x + 8, y + 10), h, fill="white", font=_font(12, bold=True))
        x += col_w[i]
    y += header_h - 10

    for idx, (nama, data) in enumerate(ringkasan.items()):
        bg = "#F1F5F9" if idx % 2 == 0 else "white"
        draw.rectangle([20, y, width - 20, y + row_h], fill=bg, outline="#E2E8F0")
        x = 20
        values = [nama, data["no_hp"], ", ".join(data["oc"]) or "-", ", ".join(data["piket"]) or "-"]
        for i, v in enumerate(values):
            draw.text((x + 8, y + 8), str(v), fill="#0F172A", font=_font(11))
            x += col_w[i]
        y += row_h

    if unmatched:
        draw.text((20, y + 6), f"⚠️ Nama gak ketemu di whitelist: {', '.join(unmatched)}",
                   fill="#B91C1C", font=_font(11))

    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)
    return send_file(output, mimetype="image/png", as_attachment=True,
                      download_name=f"jadwal_{bulan_nama}_{tahun}.png")